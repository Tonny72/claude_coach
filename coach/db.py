"""SQLite-databank voor de running-coach: één querybaar bestand (`garmin.db`) dat alle
relevante bronnen samenbrengt.

Bronnen
-------
* **Activiteiten** — FIT-bestanden (`Garmin/ACTIVITY/*.fit`, 2023→heden, met laps) en het
  handmatige Excel-logboek (2001-2025, voor dagen zonder FIT). De Garmin-Connect-historie-
  export bevat alleen geaggregeerde records en wordt daarom niet als activiteitenbron gebruikt.
* **Dagelijkse wellness** — UDS-dagsamenvattingen uit de GDPR-export (2014→2025):
  rust-HS, stress, body battery, respiratie, stappen, intensiteitsminuten, calorieën, hydratatie.
* **Referentie** — persoonlijke records en gear/materiaal (schoenkilometers).

Tabellen: ``files`` (bron-register met hash → incrementeel), ``activities``,
``activity_laps``, ``daily_wellness``, ``personal_records``, ``gear``.

Gebruik: ``python build_db.py`` (incrementeel) of ``python build_db.py --rebuild``.
"""
import datetime
import glob
import hashlib
import json
import os
import sqlite3

from . import config

SCHEMA = """
CREATE TABLE IF NOT EXISTS files (
    path        TEXT PRIMARY KEY,
    kind        TEXT,
    sha1        TEXT,
    size        INTEGER,
    mtime       REAL,
    rows        INTEGER,
    ingested_at TEXT
);

CREATE TABLE IF NOT EXISTS activities (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    date         TEXT NOT NULL,            -- YYYY-MM-DD (lokale dag)
    start_time   TEXT,                     -- ISO-datetime, NULL voor logboekrijen
    sport        TEXT,
    sub_sport    TEXT,
    name         TEXT,
    dist_km      REAL,
    dur_s        REAL,
    pace_s_km    REAL,
    avg_hr       INTEGER,
    max_hr       INTEGER,
    avg_cadence  REAL,                     -- spm, beide benen
    elev_gain_m  REAL,
    elev_loss_m  REAL,
    calories     REAL,
    aerobic_te   REAL,
    anaerobic_te REAL,
    vo2max       REAL,
    avg_stride_m REAL,
    avg_gct_ms   REAL,
    avg_vosc_cm  REAL,
    source       TEXT,                     -- 'fit' | 'gc_history' | 'log'
    source_file  TEXT
);
CREATE INDEX IF NOT EXISTS idx_act_date ON activities(date);
CREATE UNIQUE INDEX IF NOT EXISTS idx_act_start ON activities(start_time)
    WHERE start_time IS NOT NULL;

CREATE TABLE IF NOT EXISTS activity_laps (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    activity_id INTEGER REFERENCES activities(id) ON DELETE CASCADE,
    lap_index   INTEGER,
    dist_km     REAL,
    dur_s       REAL,
    pace_s_km   REAL,
    avg_hr      INTEGER,
    max_hr      INTEGER,
    avg_cadence REAL
);
CREATE INDEX IF NOT EXISTS idx_lap_act ON activity_laps(activity_id);

-- Per-seconde meetpunten (de gedetailleerde stroom uit de FIT-`record`-berichten).
-- Alleen voor hardloop-activiteiten ingelezen; voldoende voor grafieken en run-analyse.
CREATE TABLE IF NOT EXISTS activity_records (
    activity_id INTEGER REFERENCES activities(id) ON DELETE CASCADE,
    elapsed_s   INTEGER,                  -- seconden sinds start
    distance_m  REAL,
    hr          INTEGER,
    speed_ms    REAL,                     -- m/s (enhanced_speed)
    cadence     INTEGER,                  -- spm, beide benen
    altitude_m  REAL,
    power       INTEGER,                  -- watt (indien beschikbaar)
    lat         REAL,                     -- graden
    lon         REAL                      -- graden
);
CREATE INDEX IF NOT EXISTS idx_rec_act ON activity_records(activity_id, elapsed_s);

CREATE TABLE IF NOT EXISTS daily_wellness (
    date           TEXT PRIMARY KEY,       -- YYYY-MM-DD
    resting_hr     INTEGER,
    min_hr         INTEGER,
    max_hr         INTEGER,
    steps          INTEGER,
    distance_m     INTEGER,
    active_kcal    REAL,
    total_kcal     REAL,
    bmr_kcal       REAL,
    moderate_im    INTEGER,
    vigorous_im    INTEGER,
    floors_asc_m   REAL,
    stress_avg     INTEGER,
    stress_max     INTEGER,
    bb_charged     INTEGER,
    bb_drained     INTEGER,
    resp_avg       REAL,
    resp_high      REAL,
    resp_low       REAL,
    hydration_ml   REAL,
    sweat_loss_ml  REAL,
    source_file    TEXT
);

CREATE TABLE IF NOT EXISTS personal_records (
    id           INTEGER PRIMARY KEY,      -- personalRecordId
    type         TEXT,
    value        REAL,
    activity_id  INTEGER,
    created_date TEXT,
    current      INTEGER,
    source_file  TEXT
);

CREATE TABLE IF NOT EXISTS gear (
    gear_pk     INTEGER PRIMARY KEY,
    make_model  TEXT,
    status      TEXT,
    date_begin  TEXT,
    date_end    TEXT,
    max_meters  REAL,
    source_file TEXT
);
"""


# ----------------------------------------------------------------- helpers
def connect():
    con = sqlite3.connect(config.DB_PATH)
    con.execute("PRAGMA foreign_keys = ON")
    con.row_factory = sqlite3.Row
    return con


def init_schema(con):
    con.executescript(SCHEMA)
    con.commit()


def _sha1(path):
    h = hashlib.sha1()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _unchanged(con, path):
    """True als dit bestand al ongewijzigd in het register staat (zelfde sha1)."""
    row = con.execute("SELECT sha1 FROM files WHERE path=?", (str(path),)).fetchone()
    if not row:
        return False, None
    sha = _sha1(path)
    return row["sha1"] == sha, sha


def _register(con, path, kind, sha1, rows):
    st = os.stat(path)
    con.execute(
        "INSERT OR REPLACE INTO files(path, kind, sha1, size, mtime, rows, ingested_at) "
        "VALUES (?,?,?,?,?,?,?)",
        (str(path), kind, sha1, st.st_size, st.st_mtime, rows,
         datetime.datetime.now().isoformat(timespec="seconds")),
    )


def _t2s(t):
    if isinstance(t, datetime.time):
        return t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1e6
    return None


def _g(d, *keys):
    """Eerste niet-None waarde uit een dict voor de gegeven keys."""
    for k in keys:
        v = d.get(k)
        if v is not None:
            return v
    return None


# ----------------------------------------------------------------- FIT-activiteiten
_SEMI = 180.0 / 2 ** 31          # semicircles → graden


def _fit_records(ff, start_time):
    """Per-seconde meetpunten uit de FIT-`record`-berichten als rij-tuples."""
    rows = []
    for m in ff.get_messages("record"):
        r = {x.name: x.value for x in m}
        ts = r.get("timestamp")
        if ts is None:
            continue
        cad = r.get("cadence")
        if cad is not None:
            cad = round((cad + (r.get("fractional_cadence") or 0)) * 2)   # spm, beide benen
        lat, lon = r.get("position_lat"), r.get("position_long")
        rows.append((
            int((ts - start_time).total_seconds()),
            r.get("distance"),
            r.get("heart_rate"),
            _g(r, "enhanced_speed", "speed"),
            cad,
            _g(r, "enhanced_altitude", "altitude"),
            r.get("power"),
            lat * _SEMI if lat is not None else None,
            lon * _SEMI if lon is not None else None,
        ))
    return rows


def _ingest_fit(con):
    try:
        from fitparse import FitFile
    except Exception:
        return 0, 0
    files = sorted(glob.glob(str(config.ACT_DIR / "*.fit")) +
                   glob.glob(str(config.ACT_DIR / "*.FIT")))
    n_act = n_files = 0
    for f in files:
        unchanged, sha = _unchanged(con, f)
        if unchanged:
            continue
        sha = sha or _sha1(f)
        try:
            ff = FitFile(f)
            sess, laps = None, []
            for m in ff.get_messages("session"):
                sess = {x.name: x.value for x in m}
            for m in ff.get_messages("lap"):
                laps.append({x.name: x.value for x in m})
        except Exception:
            _register(con, f, "fit", sha, 0)
            continue
        # oude rijen van dit bestand opruimen (incrementele herinlees)
        con.execute("DELETE FROM activities WHERE source_file=?", (f,))
        if not sess or not sess.get("start_time"):
            _register(con, f, "fit", sha, 0)
            continue
        st = sess["start_time"]
        dist = (sess.get("total_distance") or 0) / 1000
        sec = sess.get("total_timer_time") or 0
        cad = _g(sess, "avg_running_cadence", "avg_cadence")
        cad = cad * 2 if (cad and sess.get("sport") == "running") else cad
        stance = sess.get("avg_stance_time")
        vosc = sess.get("avg_vertical_oscillation")
        cur = con.execute(
            "INSERT OR IGNORE INTO activities(date, start_time, sport, sub_sport, dist_km, "
            "dur_s, pace_s_km, avg_hr, max_hr, avg_cadence, elev_gain_m, elev_loss_m, "
            "calories, aerobic_te, anaerobic_te, avg_stride_m, avg_gct_ms, avg_vosc_cm, "
            "source, source_file) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'fit', ?)",
            (st.date().isoformat(), st.isoformat(sep=" "), sess.get("sport"),
             sess.get("sub_sport"), round(dist, 3), sec, (sec / dist) if dist else None,
             sess.get("avg_heart_rate"), sess.get("max_heart_rate"), cad,
             sess.get("total_ascent"), sess.get("total_descent"), sess.get("total_calories"),
             sess.get("total_training_effect"), sess.get("total_anaerobic_training_effect"),
             sess.get("avg_step_length"), stance, vosc, f),
        )
        if cur.rowcount:
            n_act += 1
            act_id = cur.lastrowid
            for i, l in enumerate(laps):
                ld = (l.get("total_distance") or 0) / 1000
                lt = l.get("total_timer_time") or 0
                lc = _g(l, "avg_running_cadence", "avg_cadence")
                lc = lc * 2 if (lc and sess.get("sport") == "running") else lc
                con.execute(
                    "INSERT INTO activity_laps(activity_id, lap_index, dist_km, dur_s, "
                    "pace_s_km, avg_hr, max_hr, avg_cadence) VALUES (?,?,?,?,?,?,?,?)",
                    (act_id, i + 1, round(ld, 3), lt, (lt / ld) if ld else None,
                     l.get("avg_heart_rate"), l.get("max_heart_rate"), lc),
                )
            if sess.get("sport") == "running":
                recs = _fit_records(ff, st)
                if recs:
                    con.executemany(
                        "INSERT INTO activity_records(activity_id, elapsed_s, distance_m, hr, "
                        "speed_ms, cadence, altitude_m, power, lat, lon) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?)",
                        [(act_id, *r) for r in recs],
                    )
        _register(con, f, "fit", sha, cur.rowcount)
        n_files += 1
    con.commit()
    return n_act, n_files


# NB: de Garmin-Connect-historie (`summarizedActivities.json`, 2011-2016) is bewust GEEN
# activiteitenbron: die export bevat uitsluitend geaggregeerde/gegroepeerde records
# (opgetelde afstand/duur, geen losse activiteiten). Het Excel-logboek dekt 2011-2016 schoon.


# ----------------------------------------------------------------- Excel-logboek (handmatig)
def _ingest_log(con):
    """Logboekrijen (2001-2025) voor dagen die nog niet door FIT/GC gedekt zijn."""
    if not config.XLSX.exists():
        return 0
    unchanged, sha = _unchanged(con, config.XLSX)
    if unchanged:
        return 0
    sha = sha or _sha1(config.XLSX)
    con.execute("DELETE FROM activities WHERE source='log'")
    import openpyxl
    wb = openpyxl.load_workbook(config.XLSX, data_only=True)
    ws = wb["Reeksen"]
    # dagen die al door een getimede bron gedekt zijn
    covered = {r["date"] for r in con.execute(
        "SELECT DISTINCT date FROM activities WHERE source IN ('fit','gc_history')")}
    n = 0
    for r in range(2, ws.max_row + 1):
        dt = ws.cell(r, 1).value
        if not isinstance(dt, datetime.datetime):
            continue
        d = dt.date().isoformat()
        if d in covered:
            continue
        dist = ws.cell(r, 2).value
        sec = _t2s(ws.cell(r, 3).value)
        pace = _t2s(ws.cell(r, 4).value)
        hr = ws.cell(r, 5).value
        mx = ws.cell(r, 6).value
        if not isinstance(dist, (int, float)) or not sec or dist <= 0:
            continue
        hr = hr if isinstance(hr, (int, float)) and hr <= 220 else None
        mx = mx if isinstance(mx, (int, float)) and mx <= 220 else None
        pace = pace or (sec / dist)
        if pace > 12 * 60:
            continue
        con.execute(
            "INSERT INTO activities(date, sport, dist_km, dur_s, pace_s_km, avg_hr, max_hr, "
            "source, source_file) VALUES (?, 'running', ?,?,?,?,?, 'log', ?)",
            (d, float(dist), sec, pace, hr, mx, str(config.XLSX)),
        )
        n += 1
    _register(con, config.XLSX, "log", sha, n)
    con.commit()
    return n


# ----------------------------------------------------------------- dagelijkse wellness (UDS)
def _ingest_wellness(con):
    files = sorted(glob.glob(str(config.EXPORT_DIR / "**" / "UDSFile_*.json"), recursive=True))
    n_days = n_files = 0
    for f in files:
        unchanged, sha = _unchanged(con, f)
        if unchanged:
            continue
        sha = sha or _sha1(f)
        try:
            recs = json.load(open(f, encoding="utf-8"))
        except Exception:
            _register(con, f, "uds", sha, 0)
            continue
        cnt = 0
        for rec in (recs if isinstance(recs, list) else []):
            d = rec.get("calendarDate")
            if not d:
                continue
            stress = _total_stress(rec.get("allDayStress"))
            bb = rec.get("bodyBattery") or {}
            resp = rec.get("respiration") or {}
            hyd = rec.get("hydration") or {}
            con.execute(
                "INSERT OR REPLACE INTO daily_wellness(date, resting_hr, min_hr, max_hr, steps, "
                "distance_m, active_kcal, total_kcal, bmr_kcal, moderate_im, vigorous_im, "
                "floors_asc_m, stress_avg, stress_max, bb_charged, bb_drained, resp_avg, "
                "resp_high, resp_low, hydration_ml, sweat_loss_ml, source_file) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (d, rec.get("restingHeartRate"), rec.get("minHeartRate"), rec.get("maxHeartRate"),
                 rec.get("totalSteps"), rec.get("totalDistanceMeters"),
                 rec.get("activeKilocalories"), rec.get("totalKilocalories"),
                 rec.get("bmrKilocalories"), rec.get("moderateIntensityMinutes"),
                 rec.get("vigorousIntensityMinutes"), rec.get("floorsAscendedInMeters"),
                 stress[0], stress[1], bb.get("chargedValue"), bb.get("drainedValue"),
                 resp.get("avgWakingRespirationValue"), resp.get("highestRespirationValue"),
                 resp.get("lowestRespirationValue"), hyd.get("valueInML"),
                 hyd.get("sweatLossInML"), f),
            )
            cnt += 1
        _register(con, f, "uds", sha, cnt)
        n_days += cnt
        n_files += 1
    con.commit()
    return n_days, n_files


def _total_stress(ads):
    if not isinstance(ads, dict):
        return None, None
    for agg in ads.get("aggregatorList") or []:
        if agg.get("type") == "TOTAL":
            return agg.get("averageStressLevel"), agg.get("maxStressLevel")
    return None, None


# ----------------------------------------------------------------- referentie (PR's, gear)
def _ingest_reference(con):
    base = config.EXPORT_DIR
    n_pr = n_gear = 0
    pr_files = glob.glob(str(base / "**" / "*personalRecord.json"), recursive=True)
    if pr_files:
        f = pr_files[0]
        unchanged, sha = _unchanged(con, f)
        if not unchanged:
            sha = sha or _sha1(f)
            data = json.load(open(f, encoding="utf-8"))
            d = data[0] if isinstance(data, list) else data
            for pr in d.get("personalRecords", []):
                con.execute(
                    "INSERT OR REPLACE INTO personal_records(id, type, value, activity_id, "
                    "created_date, current, source_file) VALUES (?,?,?,?,?,?,?)",
                    (pr.get("personalRecordId"), pr.get("personalRecordType"), pr.get("value"),
                     pr.get("activityId"), pr.get("createdDate"),
                     1 if pr.get("current") else 0, f),
                )
                n_pr += 1
            _register(con, f, "personal_records", sha, n_pr)
    gear_files = glob.glob(str(base / "**" / "*gear.json"), recursive=True)
    if gear_files:
        f = gear_files[0]
        unchanged, sha = _unchanged(con, f)
        if not unchanged:
            sha = sha or _sha1(f)
            data = json.load(open(f, encoding="utf-8"))
            d = data[0] if isinstance(data, list) else data
            for g in d.get("gearDTOS", []):
                con.execute(
                    "INSERT OR REPLACE INTO gear(gear_pk, make_model, status, date_begin, "
                    "date_end, max_meters, source_file) VALUES (?,?,?,?,?,?,?)",
                    (g.get("gearPk"), g.get("customMakeModel") or g.get("displayName"),
                     g.get("gearStatusName"), g.get("dateBegin"), g.get("dateEnd"),
                     g.get("maximumMeters"), f),
                )
                n_gear += 1
            _register(con, f, "gear", sha, n_gear)
    con.commit()
    return n_pr, n_gear


# ----------------------------------------------------------------- orchestratie
def build(rebuild=False):
    """Bouw/ververs de databank. `rebuild=True` gooit alles weg en leest opnieuw in."""
    if rebuild and config.DB_PATH.exists():
        config.DB_PATH.unlink()
    con = connect()
    init_schema(con)
    res = {}
    res["fit_act"], res["fit_files"] = _ingest_fit(con)
    res["log"] = _ingest_log(con)
    res["wellness_days"], res["wellness_files"] = _ingest_wellness(con)
    res["pr"], res["gear"] = _ingest_reference(con)
    con.close()
    return res


def summary():
    """Korte inhoudsopgave van de databank (voor de CLI)."""
    con = connect()
    out = []
    a = con.execute("SELECT COUNT(*) c, MIN(date) lo, MAX(date) hi FROM activities").fetchone()
    out.append(f"activities      : {a['c']:>6}  ({a['lo']} → {a['hi']})")
    for src in ("fit", "log"):
        r = con.execute("SELECT COUNT(*) c FROM activities WHERE source=?", (src,)).fetchone()
        out.append(f"   · {src:<11}: {r['c']:>6}")
    lp = con.execute("SELECT COUNT(*) c FROM activity_laps").fetchone()
    out.append(f"activity_laps   : {lp['c']:>6}")
    rc = con.execute("SELECT COUNT(*) c, COUNT(DISTINCT activity_id) a "
                     "FROM activity_records").fetchone()
    out.append(f"activity_records: {rc['c']:>6}  (per-sec stroom van {rc['a']} runs)")
    w = con.execute("SELECT COUNT(*) c, MIN(date) lo, MAX(date) hi FROM daily_wellness").fetchone()
    out.append(f"daily_wellness  : {w['c']:>6}  ({w['lo']} → {w['hi']})")
    pr = con.execute("SELECT COUNT(*) c FROM personal_records").fetchone()
    g = con.execute("SELECT COUNT(*) c FROM gear").fetchone()
    out.append(f"personal_records: {pr['c']:>6}")
    out.append(f"gear            : {g['c']:>6}")
    con.close()
    return "\n".join(out)


# ----------------------------------------------------------------- toegang (grafieken/analyse)
def find_activity(con, date, sport="running"):
    """Activity-id van de (FIT-)activiteit op `date` (YYYY-MM-DD of date), of None.
    Kiest de activiteit met de meeste detailpunten als er meerdere op één dag zijn."""
    if hasattr(date, "isoformat"):
        date = date.isoformat()
    row = con.execute(
        "SELECT a.id FROM activities a "
        "WHERE a.date=? AND (? IS NULL OR a.sport=?) AND a.source='fit' "
        "ORDER BY (SELECT COUNT(*) FROM activity_records r WHERE r.activity_id=a.id) DESC "
        "LIMIT 1", (date, sport, sport)).fetchone()
    return row["id"] if row else None


def run_stream(con, activity_id):
    """Per-seconde stroom van een run als dict van kolom-lijsten — klaar voor grafieken.

    Bevat: elapsed_s, distance_m, hr, speed_ms, pace_s_km, cadence, altitude_m, power, lat, lon.
    Pace is afgeleid van speed (None waar speed 0/ontbreekt). Lege lijsten als er geen detail is.
    """
    rows = con.execute(
        "SELECT elapsed_s, distance_m, hr, speed_ms, cadence, altitude_m, power, lat, lon "
        "FROM activity_records WHERE activity_id=? ORDER BY elapsed_s", (activity_id,)
    ).fetchall()
    cols = ("elapsed_s", "distance_m", "hr", "speed_ms", "cadence", "altitude_m",
            "power", "lat", "lon")
    out = {c: [r[c] for r in rows] for c in cols}
    out["pace_s_km"] = [(1000.0 / s) if s else None for s in out["speed_ms"]]
    return out
