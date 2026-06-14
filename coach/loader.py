"""Laadt alle lopen uit het Excel-logboek (2001-2025) + recente Fenix .fit-bestanden."""
import datetime, glob, json, os
from dataclasses import dataclass, asdict

from . import config


@dataclass
class Run:
    d: datetime.date
    dist: float          # km
    sec: float           # duur (s)
    pace: float          # s/km
    hr: float | None     # gem. HS
    mx: float | None     # max HS
    src: str             # 'log' of 'fit'


def _time_to_sec(t):
    if isinstance(t, datetime.time):
        return t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1e6
    return None


def _load_reeksen():
    import openpyxl
    wb = openpyxl.load_workbook(config.XLSX, data_only=True)
    ws = wb["Reeksen"]
    out = []
    for r in range(2, ws.max_row + 1):
        dt = ws.cell(r, 1).value
        if not isinstance(dt, datetime.datetime):
            continue
        dist = ws.cell(r, 2).value
        sec = _time_to_sec(ws.cell(r, 3).value)
        pace = _time_to_sec(ws.cell(r, 4).value)
        hr = ws.cell(r, 5).value
        mx = ws.cell(r, 6).value
        if not dist or not sec or dist <= 0:
            continue
        hr = hr if isinstance(hr, (int, float)) else None
        mx = mx if isinstance(mx, (int, float)) else None
        # ruisfilter: onmogelijke waarden weggooien
        if mx and mx > 220:
            mx = None
        if hr and hr > 220:
            hr = None
        pace = pace or sec / dist
        if pace > 12 * 60:        # >12 min/km = wandeling/meetfout
            continue
        if not isinstance(dist, (int, float)):
            continue
        out.append(Run(dt.date(), float(dist), sec, pace, hr, mx, "log"))
    return out


def _load_fit_after(cutoff: datetime.date):
    from fitparse import FitFile
    out = []
    files = sorted(glob.glob(str(config.ACT_DIR / "2025-*.fit")) +
                   glob.glob(str(config.ACT_DIR / "2026-*.fit")))
    for f in files:
        try:
            ff = FitFile(f)
        except Exception:
            continue
        s = None
        for m in ff.get_messages("session"):
            s = {x.name: x.value for x in m}
        if not s or not s.get("start_time") or s.get("sport") != "running":
            continue
        d = s["start_time"].date()
        if d <= cutoff:
            continue
        dist = (s.get("total_distance") or 0) / 1000
        sec = s.get("total_timer_time") or 0
        if dist <= 0 or sec <= 0:
            continue
        out.append(Run(d, dist, sec, sec / dist,
                       s.get("avg_heart_rate"), s.get("max_heart_rate"), "fit"))
    return out


def load_runs(use_cache=False) -> list[Run]:
    """Volledige, gesorteerde lijst van lopen uit beide bronnen."""
    if use_cache and config.CACHE_JSON.exists():
        raw = json.load(open(config.CACHE_JSON))
        return [Run(datetime.date.fromisoformat(r["d"]), r["dist"], r["sec"],
                    r["pace"], r["hr"], r["mx"], r["src"]) for r in raw]
    runs = _load_reeksen()
    last_log = max(r.d for r in runs)
    runs += _load_fit_after(last_log)
    runs.sort(key=lambda r: r.d)
    # cache
    with open(config.CACHE_JSON, "w") as fh:
        json.dump([{**asdict(r), "d": r.d.isoformat()} for r in runs], fh)
    return runs
